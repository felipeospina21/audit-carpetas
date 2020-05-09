import openpyxl as xl
import os
import getopt
import glob
import sys
from datetime import datetime
from pathlib import Path, PureWindowsPath
#from pprint import pprint
# python acuerdos.py < input_file.txt > output_file.txt

def load_acuerdos(lista_acuerdos):
    # Crea Lista con acuerdos vigentes
    os.chdir('C:/Documentos Empresa/OneDrive - MINEROS/Desktop/Automatizaciones/auditoria/acuerdos')
    sys.stdin = open("input_file.txt", 'r')
    limite = int(input())
    for i in range(1, limite-1, 1):
        lista_acuerdos.append(input())
    return lista_acuerdos

def list_oc_audit(ws, fila):
    # Crea lista con OC a auditar
    global lista_oc
    while float(ws['D{}'.format(fila)].value) >= 17556060: # <= 20 SMMLV (Poner esto dinámico)
        lista_oc.append(ws['A{}'.format(fila)].value)
        fila+=1
    return lista_oc

def dict_oc_acuerdos(ws, d_acuerdos, fila, lista_oc):
    # Crea diccionario OC: cod_prov cod_mat
    while ws['F{}'.format(fila)].value in lista_oc:
        if ws['F{}'.format(fila)].value != ws['F{}'.format(fila - 1)].value: 
            lista_concat = []
        lista_concat.append(str(ws['G{}'.format(fila)].value) + ' ' + str(ws['H{}'.format(fila)].value))
        d_acuerdos.update({ws['F{}'.format(fila)].value : lista_concat})
        fila+=1
        if fila == 99999:
            break
    return d_acuerdos

def dict_informe (ws, fila, d, d_acuerdos, lista_oc):
    # crear OC:lista
    for OC in lista_oc:
        lista_dict = []
        lista_nom_excel = []
        lista_nom_pdf = []
        acuerdo = ""
        if str(OC)[0:2] == '45':
            path_audit = str(Path('//Nasestrella/oc/{}/Nacionales/{}'.format(year, OC)))
        elif str(OC)[0:2] == '55':
            path_audit = str(Path('//Nasestrella/oc/{}/Importaciones/{}'.format(year, OC)))
        # Verifica folder y archivos
        folder = os.path.exists(path_audit)
        excel_file_list = glob.glob(path_audit + '/*.xlsx')
        pdf_file_list = glob.glob(path_audit + '/*.pdf')
        # Cuenta Excel
        if folder == True and excel_file_list != []:
            count_excel = len(glob.glob(path_audit + '/*.xlsx'))
        else:
            count_excel = 0
        # Cuenta PDF
        if folder == True and pdf_file_list != []:
            count_pdf = len(glob.glob(path_audit + '/*.pdf'))
        else:
            count_pdf = 0
        # Acuerdos
        if folder == False:
            for i in d_acuerdos.get(OC):
                if i in lista_acuerdos: 
                    acuerdo = True
                    break
                else:
                    acuerdo = False
        # Nombres Excel
        if count_excel >= 1:
            for i in excel_file_list:
                lista_nom_excel.append(i[44:])
        else:
            lista_nom_excel = ['-']
        # Nombres PDF
        if count_pdf >= 1:
            for e in pdf_file_list:
                lista_nom_pdf.append(e[44:])
        else:
            lista_nom_pdf = ['-']
        # Nombres Folders
        if folder == True:
            lista_archivos = os.listdir(path_audit)
            result = os.scandir(path_audit)
            for elements in result:
                if elements.is_dir() == False:
                    lista_archivos.remove(elements.name)
            if lista_archivos == []:
                lista_archivos = ['-']
        else:
            lista_archivos = ['-']
        # Nombres Correos

        # Crea diccionario   
        lista_dict.append(ws['B{}'.format(fila)].value)
        lista_dict.append(ws['C{}'.format(fila)].value) 
        lista_dict.append(ws['D{}'.format(fila)].value) 
        lista_dict.append(path_audit)
        lista_dict.append(acuerdo)
        lista_dict.append(count_excel)
        lista_dict.append(count_pdf)
        lista_dict.append(lista_nom_excel)
        lista_dict.append(lista_nom_pdf)
        lista_dict.append(lista_archivos)
        #lista_dict.append(lista_correos) => Nombres Correos
        d.update({OC : lista_dict})
        fila+=1
    return d

def informe_excel(ws, d, fila):
    # Imprime diccionario en excel
    for elements in d:
        ws['A{}'.format(fila)].value = elements
        ws['B{}'.format(fila)].value = d.get(elements)[0]
        ws['C{}'.format(fila)].value = d.get(elements)[1] 
        ws['D{}'.format(fila)].value = d.get(elements)[2] 
        ws['E{}'.format(fila)].value = os.path.exists(d.get(elements)[3])  
        ws['F{}'.format(fila)].value = d.get(elements)[4]
        ws['G{}'.format(fila)].value = d.get(elements)[5]
        ws['H{}'.format(fila)].value = d.get(elements)[6]
        lista_nom_excel = d.get(elements)[7]
        if lista_nom_excel != []:
            concat_excel = ', '.join(lista_nom_excel)
            ws['I{}'.format(fila)].value = concat_excel
        lista_nom_pdf = d.get(elements)[8]
        if lista_nom_pdf != []:
            concat_pdf = ', '.join(lista_nom_pdf)
            ws['J{}'.format(fila)].value = concat_pdf
        lista_folders = d.get(elements)[9]
        if lista_folders != []:
            concat_subfolders = ', '.join(lista_folders)
            ws['K{}'.format(fila)].value = concat_subfolders
        # => Nombres Correos
        fila+=1

# Variables
meses = {1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril", 5:"Mayo", 6:"Junio", 7:"Julio", 8:"Agosto", 9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre"}
now = datetime.now()
year = str(now.year)
mes = now.month - 1
mes = meses.get(mes)
path_oc = Path('C:/Documentos Empresa/OneDrive - MINEROS/Desktop/INDICADORES/{}/{}/Ts_Comprador({}).xlsx'.format(year, mes, mes))
path_informe = Path('C:/Documentos Empresa/OneDrive - MINEROS/Desktop/INDICADORES/{}/{}/informe.xlsx'.format(year, mes))
wb = xl.load_workbook(path_oc)
wb_inf = xl.load_workbook(path_informe)
ws = wb['oc']
ws_inf = wb_inf['informe']
d = {}
d_acuerdos = {}
lista_oc = []
lista_acuerdos = []
lista_concat = []

# Funciones
try:
    load_acuerdos(lista_acuerdos)
    list_oc_audit(ws, 4)
    dict_oc_acuerdos(ws, d_acuerdos, 4, lista_oc)
    d = dict_informe(ws, 4, d, d_acuerdos, lista_oc)
    #pprint(d)
    informe_excel(ws_inf, d, 2)
    wb_inf.save(path_informe)

except TypeError:
    os.chdir('C:/Documentos Empresa/OneDrive - MINEROS/Desktop/Automatizaciones/auditoria/acuerdos')
    sys.stdout = open("output_file.txt", 'w')
    print('Error iterando la lista de acuerdos del informe, revisar variable fila de la función "dict_oc_acuerdos"')
   
