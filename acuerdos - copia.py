import openpyxl as xl
import os
import getopt
import glob
import sys
from datetime import datetime
from pathlib import Path, PureWindowsPath
from pprint import pprint
from time import time
# python acuerdos.py < input_file.txt > output_file.txt
# El proceso está tomando demasiado tiempo para ejecutar, revisar como optimizar.

def list_oc_audit(ws, fila, d, lista_acuerdos):
    # Crea lista con OC a auditar
    lista_concat = []
    lista_oc =[]
    d_acuerdos = {}
    while float(ws[f'D{fila}'].value) >= 17556060: # <= 20 SMMLV (Poner esto dinámico)
        lista_oc.append(ws[f'A{fila}'].value)
        fila+=1
        if fila == 99999:
            break
    fila = 4
    limite = len(lista_oc)
    # Crea {OC : cod_prov cod_mat}
    while ws[f'F{fila}'].value in lista_oc:
        if ws[f'F{fila}'].value != ws[f'F{fila-1}'].value: 
            lista_concat = []
        lista_concat.append(str(ws[f'G{fila}'].value) + ' ' + str(ws[f'H{fila}'].value))
        d_acuerdos.update({ws[f'F{fila}'].value : lista_concat})
        fila+=1
        if fila == 99999:
            break
    fila = 4       
    # Rutas
    path_audit_nacional = str(Path(f'//Nasestrella/oc/{year}/Nacionales'))
    path_audit_internacional = str(Path(f'//Nasestrella/oc/{year}/Importaciones'))
    lista_ordenes_nacional = os.listdir(path_audit_nacional)
    lista_ordenes_internacional = os.listdir(path_audit_internacional)
    # Crear {OC : lista}
    for i in range(0, limite):
        lista_nom_excel = []
        lista_nom_pdf = []
        lista_nom_msg = []
        lista_ordenes = []
        lista_resto = []
        lista_dict = []
        folder  = False
        acuerdo = False
        if str(ws[f'A{fila}'].value)[0:2] == '45':
            path_audit = path_audit_nacional
            lista_ordenes = lista_ordenes_nacional
        elif str(ws[f'A{fila}'].value)[0:2] == '55':
            path_audit = path_audit_internacional
            lista_ordenes = lista_ordenes_internacional 
        # Verifica folder 
        for ordenes in lista_ordenes:
            if ordenes.find(str(ws[f'A{fila}'].value)) != -1:
                if os.path.isdir(f'{path_audit}/{ordenes}'):
                    folder = True
                    sub_folder = os.listdir(f'{path_audit}/{ordenes}')
                    # Valida archivos
                    for elements in sub_folder:
                        try:
                            nombre, extension = elements.split('.')
                            if extension == 'xlsx':
                                lista_nom_excel.append(elements)
                            elif extension == 'pdf':
                                lista_nom_pdf.append(elements)
                            elif extension == 'msg':
                                lista_nom_msg.append(elements)
                            else:
                                lista_resto.append(elements)
                        except ValueError:
                            lista_resto.append(elements)
                    break
        # Verifica Acuerdos
        if folder == False:
            for element in d_acuerdos.get(ws[f'A{fila}'].value):
                if element in lista_acuerdos: 
                    acuerdo = True
                    break            
        # Crea diccionario   
        lista_dict.append(ws[f'B{fila}'].value)
        lista_dict.append(ws[f'C{fila}'].value) 
        lista_dict.append(ws[f'D{fila}'].value)
        lista_dict.append(folder)
        lista_dict.append(acuerdo)
        lista_dict.append(lista_nom_excel)
        lista_dict.append(lista_nom_pdf)
        lista_dict.append(lista_nom_msg)
        lista_dict.append(lista_resto) 
        d.update({ws[f'A{fila}'].value : lista_dict})
        fila+=1
    return d

# Variables
t1 = time()
meses = {1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril", 5:"Mayo", 6:"Junio", 7:"Julio", 8:"Agosto", 9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre"}
now = datetime.now()
year = str(now.year)
mes = now.month - 1
mes = meses.get(mes)
path_oc = Path('C:/Documentos Empresa/OneDrive - MINEROS/Desktop/INDICADORES/{0}/{1}/Ts_Comprador({1}).xlsx'.format(year, mes))
path_informe = Path('C:/Documentos Empresa/OneDrive - MINEROS/Desktop/INDICADORES/{0}/{1}/informe.xlsx'.format(year, mes))
wb = xl.load_workbook(path_oc)
wb_inf = xl.load_workbook(path_informe)
ws = wb['oc']
ws_inf = wb_inf['informe']
d = {}

# Crea Lista con acuerdos vigentes
os.chdir('C:/Documentos Empresa/OneDrive - MINEROS/Desktop/Automatizaciones/auditoria/acuerdos')
sys.stdin = open("input_file.txt", 'r')
limite = int(input())
lista_acuerdos = [input() for i in range(1, limite-1)]

# Funciones
list_oc_audit(ws, 4, d, lista_acuerdos)
pprint(d)
t2 = time()
print (f'tiempo: {t2 - t1}')
