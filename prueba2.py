import openpyxl as xl
import os
import getopt
import glob
import sys
from datetime import datetime
from pathlib import Path, PureWindowsPath
from pprint import pprint
# python acuerdos.py < input_file.txt > output_file.txt

def list_oc_audit(ws, fila, d_acuerdos, lista_oc, lista_concat):
    # Crea lista con OC a auditar
    while float(ws['D{}'.format(fila)].value) >= 17556060: # <= 20 SMMLV (Poner esto dinámico)
        lista_oc.append(ws['A{}'.format(fila)].value)
        fila+=1
        if fila == 99999:
            break
    # Crea diccionario OC: cod_prov cod_mat
    while ws['F{}'.format(fila)].value in lista_oc:
        if ws['F{}'.format(fila)].value != ws['F{}'.format(fila - 1)].value: 
            lista_concat = []
        lista_concat.append(str(ws['G{}'.format(fila)].value) + ' ' + str(ws['H{}'.format(fila)].value))
        d_acuerdos.update({ws['F{}'.format(fila)].value : lista_concat})
        fila+=1
        if fila == 99999:
            break
    return d_acuerdos, lista_oc

def dict_informe (ws, fila, d, d_acuerdos, lista_oc):
    # crear OC:lista
    for OC in lista_oc:
        folder_files_list = []
        lista_ordenes = []
        lista_dict = []
        lista_nom_excel = []
        lista_nom_pdf = []
        acuerdo = ""
        folder = False
        if str(OC)[0:2] == '45':
            path_audit = str(Path('//Nasestrella/oc/{}/Nacionales'.format(year)))
        elif str(OC)[0:2] == '55':
            path_audit = str(Path('//Nasestrella/oc/{}/Importaciones'.format(year)))
        elif str(OC)[0:2] == '62':
            path_audit = str(Path('//Nasestrella/oc/{}/Servicios'.format(year)))
            
        # Verifica folder 
        lista_ordenes = os.listdir(path_audit)
        for ordenes in lista_ordenes:
	        if ordenes.find(OC) != -1:
		        folder = True
		        # Verifica Archivos
                folder_files_list = os.listdir('{}/{}'.format(path_audit, ordenes))
            else:
                for element in d_acuerdos.get(OC):
                    if element in lista_acuerdos: 
                        acuerdo = True
                        break
                    else:
                        acuerdo = False                
        
        

        # excel_file_list = os.listdir('{}/{}/*.xlsx'.format(path_audit, OC))
        # pdf_file_list = glob.glob('{}/{}/*.pdf'.format(path_audit, OC))
        # Cuenta Excel
#         if folder == True and excel_file_list != []:
#             count_excel = len(glob.glob(path_audit + '/*.xlsx'))
#         else:
#             count_excel = 0
#         # Cuenta PDF
#         if folder == True and pdf_file_list != []:
#             count_pdf = len(glob.glob(path_audit + '/*.pdf'))
#         else:
#             count_pdf = 0
# x
#         # Nombres Excel
#         if count_excel >= 1:
#             for i in excel_file_list:
#                 lista_nom_excel.append(i[44:])
#         else:
#             lista_nom_excel = ['-']
#         # Nombres PDF
#         if count_pdf >= 1:
#             for e in pdf_file_list:
#                 lista_nom_pdf.append(e[44:])
#         else:
#             lista_nom_pdf = ['-']
#         # Nombres Folders
#         if folder == True:
#             lista_archivos = os.listdir(path_audit)
#             result = os.scandir(path_audit)
#             for elements in result:
#                 if elements.is_dir() == False:
#                     lista_archivos.remove(elements.name)
#             if lista_archivos == []:
#                 lista_archivos = ['-']
#         else:
#             lista_archivos = ['-']
        # Nombres Correos

        # Crea diccionario   
        lista_dict.append(ws['B{}'.format(fila)].value)
        lista_dict.append(ws['C{}'.format(fila)].value) 
        lista_dict.append(ws['D{}'.format(fila)].value) 
        lista_dict.append(path_audit)
        lista_dict.append(acuerdo)
        # lista_dict.append(count_excel)
        # lista_dict.append(count_pdf)
        # lista_dict.append(lista_nom_excel)
        # lista_dict.append(lista_nom_pdf)
        # lista_dict.append(lista_archivos)
        #lista_dict.append(lista_correos) => Nombres Correos
        d.update({OC : lista_dict})
        fila+=1
    return d

# Variables
meses = {1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril", 5:"Mayo", 6:"Junio", 7:"Julio", 8:"Agosto", 9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre"}
now = datetime.now()
year = str(now.year)
mes = now.month - 1
mes = meses.get(mes)
path_oc = Path('C:/Documentos Empresa/OneDrive - MINEROS/Desktop/INDICADORES/{}/{}/Ts_Comprador({}).xlsx'.format(year, mes, mes))
path_informe = Path('C:/Documentos Empresa/OneDrive - MINEROS/Desktop/INDICADORES/{}/{}/informe.xlsx'.format(year, mes))
wb = xl.load_workbook(path_oc)
wb_inf = xl.load_workbook(path_informe)
ws = wb['oc']
ws_inf = wb_inf['informe']
d = {}
d_acuerdos = {}
lista_oc = []
lista_acuerdos = []
lista_concat = []

# Crea Lista con acuerdos vigentes
os.chdir('C:/Documentos Empresa/OneDrive - MINEROS/Desktop/Automatizaciones/auditoria/acuerdos')
sys.stdin = open("input_file.txt", 'r')
limite = int(input())
lista_acuerdos = [input() for i in range(1, limite-1)]

# Funciones
list_oc_audit(ws, 4, d_acuerdos, lista_oc, lista_concat)
dict_informe(ws, 4, d, d_acuerdos, lista_oc)
print(d)

